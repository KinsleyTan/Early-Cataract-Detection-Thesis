"""Fixed-split metadata loading and deterministic TensorFlow input pipelines."""

from __future__ import annotations

from collections import Counter
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from utils import dataset_root


REQUIRED_COLUMNS = (
    "id",
    "filename",
    "eye_side",
    "diagnosis",
    "cataract_grade",
    "slit_lamp_illumination_type",
)


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@dataclass(frozen=True)
class MetadataRow:
    split: str
    excel_row: int
    subject_id: str
    filename: str
    eye_side: str
    diagnosis: str
    cataract_grade: str
    illumination_type: str
    cataract_type: str
    image_quality: str


@dataclass(frozen=True)
class Sample:
    split: str
    excel_row: int
    image_path: str
    filename: str
    subject_id: str
    eye_side: str
    diagnosis: str
    cataract_grade: str
    illumination_type: str
    cataract_type: str
    image_quality: str
    label: int

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def load_metadata(config: dict[str, Any], split: str) -> list[MetadataRow]:
    split_cfg = config["fixed_splits"][split]
    workbook_path = dataset_root(config) / split_cfg["workbook"]
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    raw_headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [clean(value) for value in raw_headers]
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        workbook.close()
        raise ValueError(f"{split} is missing required columns: {missing}")

    rows: list[MetadataRow] = []
    for excel_row, values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        cells = [clean(value) for value in values]
        if not any(cells):
            continue
        row = {
            header: cells[index] if index < len(cells) else ""
            for index, header in enumerate(headers)
            if header
        }
        rows.append(
            MetadataRow(
                split=split,
                excel_row=excel_row,
                subject_id=row.get("id", ""),
                filename=row.get("filename", ""),
                eye_side=row.get("eye_side", ""),
                diagnosis=row.get("diagnosis", ""),
                cataract_grade=row.get("cataract_grade", ""),
                illumination_type=row.get("slit_lamp_illumination_type", ""),
                cataract_type=row.get("catataract_type", ""),
                image_quality=row.get("image_quality", ""),
            )
        )
    workbook.close()
    return rows


def select_samples(
    config: dict[str, Any], rows: Iterable[MetadataRow]
) -> list[Sample]:
    """Apply the modular label policy without changing source metadata."""
    policy = config["label_policy"]
    class_to_label = {
        name.casefold(): int(label) for name, label in policy["class_to_label"].items()
    }
    excluded = {name.casefold() for name in policy.get("excluded_diagnoses", [])}
    allowed_cataract_grades = policy.get("cataract_grades")
    if allowed_cataract_grades is not None:
        allowed_cataract_grades = {
            str(value).strip().casefold() for value in allowed_cataract_grades
        }

    images_dir = dataset_root(config) / "Clean images"
    samples: list[Sample] = []
    for row in rows:
        diagnosis = row.diagnosis.casefold()
        if diagnosis in excluded:
            continue
        if diagnosis not in class_to_label:
            continue
        if (
            diagnosis == "cataract"
            and allowed_cataract_grades is not None
            and row.cataract_grade.casefold() not in allowed_cataract_grades
        ):
            continue
        image_path = images_dir / row.filename
        if not image_path.is_file():
            raise FileNotFoundError(
                f"Missing image for {row.split} row {row.excel_row}: {image_path}"
            )
        samples.append(
            Sample(
                split=row.split,
                excel_row=row.excel_row,
                image_path=str(image_path.resolve()),
                filename=row.filename,
                subject_id=row.subject_id,
                eye_side=row.eye_side,
                diagnosis=row.diagnosis,
                cataract_grade=row.cataract_grade,
                illumination_type=row.illumination_type,
                cataract_type=row.cataract_type,
                image_quality=row.image_quality,
                label=class_to_label[diagnosis],
            )
        )
    return samples


def label_counts(samples: Iterable[Sample]) -> Counter[int]:
    return Counter(sample.label for sample in samples)


def build_dataset(
    samples: list[Sample],
    config: dict[str, Any],
    *,
    training: bool,
):
    """Build a dataset yielding RGB float32 images in the [0, 255] range.

    Keras EfficientNetB0 contains its own Rescaling(1/255) layer. No
    preprocess_input or external normalization is applied here.
    """
    import tensorflow as tf

    if not samples:
        raise ValueError("Cannot build a dataset from zero samples")
    image_size = tuple(int(value) for value in config["data"]["image_size"])
    batch_size = int(config["data"]["batch_size"])
    seed = int(config["experiment"]["seed"])

    paths = [sample.image_path for sample in samples]
    labels = [sample.label for sample in samples]
    data = tf.data.Dataset.from_tensor_slices((paths, labels))
    options = tf.data.Options()
    options.experimental_deterministic = True
    data = data.with_options(options)
    if training:
        data = data.shuffle(
            buffer_size=len(samples), seed=seed, reshuffle_each_iteration=True
        )

    def decode(path, label):
        content = tf.io.read_file(path)
        image = tf.io.decode_jpeg(content, channels=3)
        image = tf.image.resize(image, image_size, method="bilinear", antialias=True)
        image = tf.clip_by_value(tf.cast(image, tf.float32), 0.0, 255.0)
        image = tf.ensure_shape(image, (image_size[0], image_size[1], 3))
        return image, tf.cast(label, tf.float32)

    data = data.map(decode, num_parallel_calls=tf.data.AUTOTUNE, deterministic=True)
    if config["data"].get("cache_in_memory", False):
        data = data.cache()
    data = data.batch(batch_size, drop_remainder=False)
    return data.prefetch(tf.data.AUTOTUNE)

