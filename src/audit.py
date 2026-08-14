#!/usr/bin/env python3
"""Read-only audit of the fixed cataract dataset splits.

This script never edits, moves, deletes, relabels, or resplits dataset files.
It writes only text/CSV reports under this project unless --output-dir is used.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import math
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from PIL import Image, UnidentifiedImageError


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DATASET_ROOT = PROJECT_ROOT.parent / "Fixed Dataset" / "Clean"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "outputs" / "reports"

# The official split assignments are fixed. This script does not create a split.
SPLIT_FILES = {
    "train": "train.xlsx",
    "validation": "val.xlsx",
    "test": "test.xlsx",
}

REQUIRED_COLUMNS = (
    "id",
    "filename",
    "eye_side",
    "diagnosis",
    "cataract_grade",
    "reflection",
    "image_quality",
    "slit_lamp_illumination_type",
    "catataract_type",  # Source spelling is intentionally preserved.
)

# Change only this policy when the future thesis label definition is finalized.
LABEL_POLICY = {
    "normal": 0,
    "cataract": 1,
    "other": None,
}

# Dataset-specific metadata vocabulary. Comparisons are case-insensitive.
KNOWN_CATARACT_GRADES = {"not applicable", "mild", "severe"}
CATARACT_GRADES = {"mild", "severe"}
KNOWN_EYE_SIDES = {"left", "right"}


def text(value: Any) -> str:
    """Convert Excel values to stable, stripped strings."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def canonical(value: Any) -> str:
    return text(value).casefold()


def display(value: str) -> str:
    return value if value else "<MISSING>"


def compact(values: Iterable[str], limit: int = 30) -> str:
    items = sorted({str(v) for v in values}, key=str.casefold)
    if len(items) <= limit:
        return ", ".join(items) if items else "none"
    return ", ".join(items[:limit]) + f", ... (+{len(items) - limit} more)"


@dataclass
class Record:
    split: str
    excel_row: int
    values: dict[str, str]
    image_path: Path | None = None

    @property
    def subject_id(self) -> str:
        return self.values.get("id", "")

    @property
    def filename(self) -> str:
        return self.values.get("filename", "")

    @property
    def diagnosis(self) -> str:
        return self.values.get("diagnosis", "")

    def locator(self) -> str:
        return f"{self.split}!row {self.excel_row}"

    def identity(self) -> str:
        return (
            f"{self.locator()}: filename={display(self.filename)}, "
            f"id={display(self.subject_id)}, diagnosis={display(self.diagnosis)}"
        )


@dataclass
class Issue:
    severity: str
    category: str
    split: str
    message: str


@dataclass
class WorkbookInfo:
    split: str
    path: Path
    sheet_name: str
    headers: list[str]
    blank_rows: list[int] = field(default_factory=list)


class CsvReport:
    def __init__(self) -> None:
        self.rows: list[dict[str, Any]] = []

    def add(
        self,
        section: str,
        severity: str,
        split: str,
        item: str,
        count: int | str = "",
        details: str = "",
    ) -> None:
        self.rows.append(
            {
                "section": section,
                "severity": severity,
                "split": split,
                "item": item,
                "count": count,
                "details": details,
            }
        )

    def write(self, path: Path) -> None:
        with path.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=("section", "severity", "split", "item", "count", "details"),
            )
            writer.writeheader()
            writer.writerows(self.rows)


def load_split(split: str, path: Path) -> tuple[WorkbookInfo, list[Record]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if not workbook.sheetnames:
        raise ValueError(f"Workbook has no worksheets: {path}")
    worksheet = workbook[workbook.sheetnames[0]]
    raw_headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [text(value) for value in raw_headers]
    records: list[Record] = []
    blank_rows: list[int] = []

    for excel_row, raw_row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        values = [text(value) for value in raw_row]
        if not any(values):
            blank_rows.append(excel_row)
            continue
        row_dict = {
            header: values[index] if index < len(values) else ""
            for index, header in enumerate(headers)
            if header
        }
        records.append(Record(split=split, excel_row=excel_row, values=row_dict))

    workbook.close()
    return (
        WorkbookInfo(
            split=split,
            path=path,
            sheet_name=worksheet.title,
            headers=headers,
            blank_rows=blank_rows,
        ),
        records,
    )


def build_image_index(images_dir: Path) -> tuple[list[Path], dict[str, list[Path]]]:
    files = sorted((path for path in images_dir.rglob("*") if path.is_file()), key=str)
    by_name: dict[str, list[Path]] = defaultdict(list)
    for path in files:
        by_name[path.name.casefold()].append(path)
    return files, by_name


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def perceptual_hash(path: Path, hash_size: int = 8, image_size: int = 32) -> int:
    """Return a 64-bit pHash using a low-frequency 2-D DCT.

    The implementation uses Pillow plus the Python standard library so the audit
    does not depend on an additional image-hashing package.
    """
    with Image.open(path) as image:
        gray = image.convert("L").resize((image_size, image_size), Image.Resampling.LANCZOS)
        # L-mode images are one byte per pixel; tobytes() avoids Pillow's
        # deprecated Image.getdata() API while keeping the hash deterministic.
        pixels = list(gray.tobytes())

    cosine = [
        [math.cos((2 * x + 1) * u * math.pi / (2 * image_size)) for x in range(image_size)]
        for u in range(hash_size)
    ]
    coefficients: list[float] = []
    for v in range(hash_size):
        alpha_v = 1 / math.sqrt(2) if v == 0 else 1.0
        for u in range(hash_size):
            alpha_u = 1 / math.sqrt(2) if u == 0 else 1.0
            total = 0.0
            for y in range(image_size):
                row_offset = y * image_size
                cos_y = cosine[v][y]
                for x in range(image_size):
                    total += pixels[row_offset + x] * cosine[u][x] * cos_y
            coefficients.append(0.25 * alpha_u * alpha_v * total)

    threshold = statistics.median(coefficients[1:])
    result = 0
    for coefficient in coefficients:
        result = (result << 1) | int(coefficient > threshold)
    return result


def hamming_distance(first: int, second: int) -> int:
    return (first ^ second).bit_count()


def distribution(records: Iterable[Record], column: str) -> Counter[str]:
    return Counter(display(record.values.get(column, "")) for record in records)


def format_distribution(counter: Counter[str]) -> str:
    return ", ".join(
        f"{key}={value}" for key, value in sorted(counter.items(), key=lambda item: item[0].casefold())
    )


def format_group(records: Iterable[Record]) -> str:
    return " | ".join(record.identity() for record in records)


def audit(dataset_root: Path, output_dir: Path, phash_threshold: int) -> tuple[Path, Path, str]:
    images_dir = dataset_root / "Clean images"
    output_dir.mkdir(parents=True, exist_ok=True)

    issues: list[Issue] = []
    workbook_info: dict[str, WorkbookInfo] = {}
    records_by_split: dict[str, list[Record]] = {}
    all_records: list[Record] = []

    for split, filename in SPLIT_FILES.items():
        path = dataset_root / filename
        info, records = load_split(split, path)
        workbook_info[split] = info
        records_by_split[split] = records
        all_records.extend(records)

        missing_columns = [column for column in REQUIRED_COLUMNS if column not in info.headers]
        unexpected_columns = [column for column in info.headers if column not in REQUIRED_COLUMNS]
        if missing_columns:
            issues.append(Issue("ERROR", "missing columns", split, compact(missing_columns)))
        if unexpected_columns:
            issues.append(Issue("WARNING", "unexpected columns", split, compact(unexpected_columns)))
        if info.blank_rows:
            issues.append(
                Issue("WARNING", "blank spreadsheet rows", split, compact(map(str, info.blank_rows)))
            )

    image_files, image_index = build_image_index(images_dir)
    referenced_paths: set[Path] = set()

    for record in all_records:
        if not record.filename:
            continue
        matches = image_index.get(Path(record.filename).name.casefold(), [])
        if not matches:
            issues.append(Issue("ERROR", "missing image", record.split, record.identity()))
        elif len(matches) > 1:
            issues.append(
                Issue(
                    "ERROR",
                    "ambiguous image filename",
                    record.split,
                    f"{record.identity()} matches {compact(str(path) for path in matches)}",
                )
            )
        else:
            record.image_path = matches[0]
            referenced_paths.add(matches[0].resolve())
            if matches[0].name != Path(record.filename).name:
                issues.append(
                    Issue(
                        "WARNING",
                        "filename case mismatch",
                        record.split,
                        f"{record.identity()} actual={matches[0].name}",
                    )
                )

    # Essential and categorical metadata checks.
    for record in all_records:
        for column in ("diagnosis", "filename", "id"):
            if not record.values.get(column, ""):
                issues.append(
                    Issue("ERROR", f"missing {column}", record.split, record.identity())
                )

        diagnosis = canonical(record.diagnosis)
        grade = canonical(record.values.get("cataract_grade", ""))
        eye_side = canonical(record.values.get("eye_side", ""))

        if diagnosis and diagnosis not in LABEL_POLICY:
            issues.append(Issue("ERROR", "unknown diagnosis", record.split, record.identity()))
        if grade and grade not in KNOWN_CATARACT_GRADES:
            issues.append(
                Issue(
                    "WARNING",
                    "unknown cataract grade",
                    record.split,
                    f"{record.identity()}, grade={record.values.get('cataract_grade', '')}",
                )
            )
        if diagnosis == "cataract" and grade not in CATARACT_GRADES:
            issues.append(
                Issue(
                    "ERROR",
                    "inconsistent diagnosis/grade",
                    record.split,
                    f"{record.identity()}, grade={display(record.values.get('cataract_grade', ''))}",
                )
            )
        if diagnosis in {"normal", "other"} and grade != "not applicable":
            issues.append(
                Issue(
                    "ERROR",
                    "inconsistent diagnosis/grade",
                    record.split,
                    f"{record.identity()}, grade={display(record.values.get('cataract_grade', ''))}",
                )
            )
        if eye_side and eye_side not in KNOWN_EYE_SIDES:
            issues.append(
                Issue(
                    "WARNING",
                    "unknown eye side",
                    record.split,
                    f"{record.identity()}, eye_side={record.values.get('eye_side', '')}",
                )
            )

    # Exact duplicate spreadsheet rows, evaluated within each official split.
    duplicate_row_groups: list[tuple[str, list[Record]]] = []
    for split, records in records_by_split.items():
        row_groups: dict[tuple[str, ...], list[Record]] = defaultdict(list)
        headers = workbook_info[split].headers
        for record in records:
            row_groups[tuple(record.values.get(header, "") for header in headers)].append(record)
        for group in row_groups.values():
            if len(group) > 1:
                duplicate_row_groups.append((split, group))
                issues.append(
                    Issue("WARNING", "duplicate spreadsheet rows", split, format_group(group))
                )

    # Subject overlap across fixed splits.
    subject_sets = {
        split: {record.subject_id for record in records if record.subject_id}
        for split, records in records_by_split.items()
    }
    leakage_pairs: list[tuple[str, str, set[str]]] = []
    for first, second in (("train", "validation"), ("train", "test"), ("validation", "test")):
        overlap = subject_sets[first] & subject_sets[second]
        leakage_pairs.append((first, second, overlap))
        if overlap:
            issues.append(
                Issue(
                    "CRITICAL",
                    "subject leakage",
                    f"{first} vs {second}",
                    compact(overlap),
                )
            )

    # Filename duplication is reported independently of image content.
    filename_groups: dict[str, list[Record]] = defaultdict(list)
    for record in all_records:
        if record.filename:
            filename_groups[record.filename.casefold()].append(record)
    duplicate_filename_groups = [group for group in filename_groups.values() if len(group) > 1]
    for group in duplicate_filename_groups:
        group_splits = {record.split for record in group}
        issues.append(
            Issue(
                "CRITICAL" if len(group_splits) > 1 else "WARNING",
                "duplicate filename across splits" if len(group_splits) > 1 else "duplicate filename within split",
                ", ".join(sorted(group_splits)),
                format_group(group),
            )
        )

    # Cryptographic and perceptual image hashes are computed once per physical file.
    records_for_path: dict[Path, list[Record]] = defaultdict(list)
    for record in all_records:
        if record.image_path:
            records_for_path[record.image_path.resolve()].append(record)

    sha_by_path: dict[Path, str] = {}
    phash_by_path: dict[Path, int] = {}
    unreadable_paths: dict[Path, str] = {}
    for path in sorted(records_for_path, key=str):
        try:
            sha_by_path[path] = sha256_file(path)
            phash_by_path[path] = perceptual_hash(path)
        except (OSError, UnidentifiedImageError, ValueError) as exc:
            unreadable_paths[path] = f"{type(exc).__name__}: {exc}"
            for record in records_for_path[path]:
                issues.append(
                    Issue(
                        "ERROR",
                        "unreadable image",
                        record.split,
                        f"{record.identity()}, error={unreadable_paths[path]}",
                    )
                )

    sha_groups: dict[str, list[Path]] = defaultdict(list)
    for path, digest in sha_by_path.items():
        sha_groups[digest].append(path)
    exact_duplicate_groups = [
        paths
        for paths in sha_groups.values()
        if len(paths) > 1 or sum(len(records_for_path[path]) for path in paths) > 1
    ]
    exact_cross_split_groups: list[list[Path]] = []
    for paths in exact_duplicate_groups:
        splits = {record.split for path in paths for record in records_for_path[path]}
        if len(splits) > 1:
            exact_cross_split_groups.append(paths)
            issues.append(
                Issue(
                    "CRITICAL",
                    "exact image duplicate across splits",
                    ", ".join(sorted(splits)),
                    format_group(record for path in paths for record in records_for_path[path]),
                )
            )

    near_duplicate_pairs: list[tuple[Path, Path, int]] = []
    hashed_paths = sorted(phash_by_path, key=str)
    for index, first_path in enumerate(hashed_paths):
        first_splits = {record.split for record in records_for_path[first_path]}
        for second_path in hashed_paths[index + 1 :]:
            second_splits = {record.split for record in records_for_path[second_path]}
            if not first_splits.isdisjoint(second_splits):
                continue
            if sha_by_path[first_path] == sha_by_path[second_path]:
                continue
            distance = hamming_distance(phash_by_path[first_path], phash_by_path[second_path])
            if distance <= phash_threshold:
                near_duplicate_pairs.append((first_path, second_path, distance))
                issues.append(
                    Issue(
                        "WARNING",
                        "near duplicate across splits",
                        f"{compact(first_splits)} vs {compact(second_splits)}",
                        f"pHash distance={distance}; "
                        + format_group(records_for_path[first_path] + records_for_path[second_path]),
                    )
                )

    # Same-subject descriptive checks within each fixed split.
    subject_descriptions: dict[str, dict[str, list[str]]] = {}
    for split, records in records_by_split.items():
        subject_records: dict[str, list[Record]] = defaultdict(list)
        for record in records:
            if record.subject_id:
                subject_records[record.subject_id].append(record)
        multiple_images = [subject for subject, group in subject_records.items() if len(group) > 1]
        both_eyes = [
            subject
            for subject, group in subject_records.items()
            if {canonical(row.values.get("eye_side", "")) for row in group} >= {"left", "right"}
        ]
        multiple_illumination = [
            subject
            for subject, group in subject_records.items()
            if len(
                {
                    canonical(row.values.get("slit_lamp_illumination_type", ""))
                    for row in group
                    if row.values.get("slit_lamp_illumination_type", "")
                }
            )
            > 1
        ]
        subject_descriptions[split] = {
            "multiple_images": multiple_images,
            "both_eyes": both_eyes,
            "multiple_illumination": multiple_illumination,
        }

    referenced_names = {path.resolve() for path in records_for_path}
    orphan_images = [path for path in image_files if path.resolve() not in referenced_names]
    unreadable_set = set(unreadable_paths)

    # Build machine-readable report rows.
    csv_report = CsvReport()
    for split, records in records_by_split.items():
        usable = Counter(
            canonical(record.diagnosis)
            for record in records
            if canonical(record.diagnosis) in {"normal", "cataract"}
            and record.image_path
            and record.image_path.resolve() not in unreadable_set
        )
        csv_report.add("B. Split summary", "INFO", split, "total rows", len(records))
        csv_report.add("B. Split summary", "INFO", split, "unique subject IDs", len(subject_sets[split]))
        csv_report.add(
            "B. Split summary",
            "INFO",
            split,
            "rows with resolved image",
            sum(record.image_path is not None for record in records),
        )
        csv_report.add("B. Split summary", "INFO", split, "usable Normal", usable["normal"])
        csv_report.add("B. Split summary", "INFO", split, "usable Cataract", usable["cataract"])
        for column in (
            "diagnosis",
            "cataract_grade",
            "eye_side",
            "slit_lamp_illumination_type",
            "catataract_type",
            "image_quality",
        ):
            for value, count in sorted(distribution(records, column).items()):
                csv_report.add("F. Distributions", "INFO", split, column, count, value)

    for first, second, overlap in leakage_pairs:
        csv_report.add(
            "C. Subject leakage",
            "CRITICAL" if overlap else "PASS",
            f"{first} vs {second}",
            "overlapping subject IDs",
            len(overlap),
            compact(overlap),
        )
    for issue in issues:
        csv_report.add("G. Findings", issue.severity, issue.split, issue.category, 1, issue.message)

    # Determine a conservative readiness verdict.
    critical_issues = [issue for issue in issues if issue.severity == "CRITICAL"]
    error_issues = [issue for issue in issues if issue.severity == "ERROR"]
    warning_issues = [issue for issue in issues if issue.severity == "WARNING"]
    if critical_issues or error_issues:
        verdict = "NOT READY"
        verdict_reason = (
            f"Found {len(critical_issues)} critical and {len(error_issues)} error finding(s). "
            "Resolve or explicitly adjudicate these before baseline training."
        )
    elif warning_issues:
        verdict = "READY WITH WARNINGS"
        verdict_reason = (
            f"No critical/error findings, but {len(warning_issues)} warning finding(s) require review."
        )
    else:
        verdict = "READY FOR BASELINE TRAINING"
        verdict_reason = "No critical, error, or warning findings were detected by this audit."
    csv_report.add("H. Verdict", "INFO", "all", verdict, "", verdict_reason)

    lines: list[str] = []
    lines.append("CATARACT DATASET AUDIT")
    lines.append("=" * 80)
    lines.append(f"Audit time (UTC): {datetime.now(timezone.utc).isoformat()}")
    lines.append(f"Dataset root (read-only): {dataset_root.resolve()}")
    lines.append(f"Image directory: {images_dir.resolve()}")
    lines.append("Official split policy: train.xlsx, val.xlsx, and test.xlsx are fixed splits.")
    lines.append("No random split, reshuffling, cleaning, deletion, moving, or relabeling was performed.")
    lines.append("Baseline label policy: Normal -> 0; Cataract -> 1; Other -> excluded.")
    lines.append(
        "Usable means a Normal/Cataract row with a uniquely resolved, readable image; "
        "suspicious rows remain reported and are not silently discarded."
    )
    lines.append(
        f"Near-duplicate criterion: 64-bit pHash, cross-split Hamming distance <= {phash_threshold}; "
        "SHA-256-identical pairs are reported separately."
    )

    lines.extend(["", "A. DATASET STRUCTURE", "-" * 80])
    for split in SPLIT_FILES:
        info = workbook_info[split]
        lines.append(
            f"{split}: {info.path.name}; sheet={info.sheet_name}; data_rows={len(records_by_split[split])}; "
            f"columns={len(info.headers)}"
        )
        lines.append(f"  metadata: {', '.join(info.headers)}")
    extension_counts = Counter(path.suffix.casefold() or "<no extension>" for path in image_files)
    lines.append(f"Clean images inventory: {len(image_files)} file(s); {format_distribution(extension_counts)}")
    lines.append(f"Referenced physical images: {len(referenced_names)}")
    lines.append(f"Unreferenced/orphan images: {len(orphan_images)}")
    if orphan_images:
        lines.append(f"  {compact(path.name for path in orphan_images)}")

    lines.extend(["", "B. SPLIT SUMMARY", "-" * 80])
    lines.append(
        "split       rows  unique_subjects  resolved_images  usable_normal  usable_cataract  excluded_other"
    )
    for split, records in records_by_split.items():
        usable = Counter(
            canonical(record.diagnosis)
            for record in records
            if canonical(record.diagnosis) in {"normal", "cataract"}
            and record.image_path
            and record.image_path.resolve() not in unreadable_set
        )
        excluded_other = sum(canonical(record.diagnosis) == "other" for record in records)
        lines.append(
            f"{split:<11} {len(records):>4} {len(subject_sets[split]):>16} "
            f"{sum(record.image_path is not None for record in records):>16} "
            f"{usable['normal']:>14} {usable['cataract']:>16} {excluded_other:>15}"
        )
    lines.append("")
    lines.append("Within-split subject multiplicity (descriptive, not automatically leakage):")
    for split, groups in subject_descriptions.items():
        lines.append(
            f"{split}: multiple_images={len(groups['multiple_images'])} "
            f"[{compact(groups['multiple_images'])}]; both_eyes={len(groups['both_eyes'])} "
            f"[{compact(groups['both_eyes'])}]; multiple_illumination={len(groups['multiple_illumination'])} "
            f"[{compact(groups['multiple_illumination'])}]"
        )

    lines.extend(["", "C. SUBJECT LEAKAGE", "-" * 80])
    for first, second, overlap in leakage_pairs:
        status = "CRITICAL LEAKAGE ISSUE" if overlap else "PASS"
        lines.append(f"{first} vs {second}: {status}; overlap={len(overlap)}; IDs={compact(overlap)}")

    lines.extend(["", "D. EXACT DUPLICATES", "-" * 80])
    lines.append(f"Repeated filename groups (any split): {len(duplicate_filename_groups)}")
    for group in duplicate_filename_groups:
        lines.append(f"  {format_group(group)}")
    lines.append(
        f"SHA-256 duplicate-content groups (different files or repeated references): "
        f"{len(exact_duplicate_groups)}"
    )
    lines.append(f"SHA-256 duplicate-content groups crossing splits: {len(exact_cross_split_groups)}")
    for paths in exact_duplicate_groups:
        digest = sha_by_path[paths[0]]
        group_records = [record for path in paths for record in records_for_path[path]]
        lines.append(f"  sha256={digest}; {format_group(group_records)}")

    lines.extend(["", "E. NEAR DUPLICATES", "-" * 80])
    lines.append(
        f"Criterion: 64-bit pHash Hamming distance <= {phash_threshold}, only across different splits, "
        "excluding SHA-256 exact matches."
    )
    lines.append(f"Suspicious cross-split pairs: {len(near_duplicate_pairs)}")
    for first_path, second_path, distance in sorted(near_duplicate_pairs, key=lambda item: item[2]):
        lines.append(
            f"  distance={distance}; "
            f"{format_group(records_for_path[first_path] + records_for_path[second_path])}"
        )

    lines.extend(["", "F. LABEL DISTRIBUTION", "-" * 80])
    for split, records in records_by_split.items():
        lines.append(f"{split} diagnosis: {format_distribution(distribution(records, 'diagnosis'))}")
        lines.append(f"{split} cataract_grade: {format_distribution(distribution(records, 'cataract_grade'))}")
    lines.append(f"all diagnosis: {format_distribution(distribution(all_records, 'diagnosis'))}")
    lines.append(f"all cataract_grade: {format_distribution(distribution(all_records, 'cataract_grade'))}")

    lines.extend(["", "G. METADATA ISSUES AND OTHER DISTRIBUTIONS", "-" * 80])
    for split, records in records_by_split.items():
        lines.append(f"{split} eye_side: {format_distribution(distribution(records, 'eye_side'))}")
        lines.append(
            f"{split} slit_lamp_illumination_type: "
            f"{format_distribution(distribution(records, 'slit_lamp_illumination_type'))}"
        )
        lines.append(
            f"{split} catataract_type: {format_distribution(distribution(records, 'catataract_type'))}"
        )
        if "image_quality" in workbook_info[split].headers:
            lines.append(
                f"{split} image_quality: {format_distribution(distribution(records, 'image_quality'))}"
            )
        else:
            lines.append(f"{split} image_quality: <COLUMN NOT AVAILABLE>")
    lines.append("")
    lines.append(f"Duplicate spreadsheet row groups: {len(duplicate_row_groups)}")
    lines.append(f"Unreadable referenced images: {len(unreadable_paths)}")
    if not issues:
        lines.append("No metadata or integrity findings.")
    else:
        issue_counts = Counter((issue.severity, issue.category) for issue in issues)
        lines.append("Finding counts:")
        for (severity, category), count in sorted(issue_counts.items()):
            lines.append(f"  {severity}: {category} = {count}")
        lines.append("Finding details:")
        for issue in issues:
            lines.append(f"  [{issue.severity}] [{issue.split}] {issue.category}: {issue.message}")

    lines.extend(["", "H. BASELINE READINESS VERDICT", "-" * 80])
    lines.append(verdict)
    lines.append(verdict_reason)
    lines.append("The test split remains audit-only and was not used for model-development decisions.")

    text_path = output_dir / "dataset_audit.txt"
    csv_path = output_dir / "dataset_audit.csv"
    text_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    csv_report.write(csv_path)
    return text_path, csv_path, verdict


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--dataset-root",
        type=Path,
        default=DEFAULT_DATASET_ROOT,
        help="Directory containing train.xlsx, val.xlsx, test.xlsx, and Clean images/.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Directory for dataset_audit.txt and dataset_audit.csv.",
    )
    parser.add_argument(
        "--phash-threshold",
        type=int,
        default=6,
        help="Maximum 64-bit pHash Hamming distance for a suspicious cross-split pair.",
    )
    args = parser.parse_args()
    if not 0 <= args.phash_threshold <= 64:
        parser.error("--phash-threshold must be between 0 and 64")
    return args


def validate_layout(dataset_root: Path) -> None:
    missing = [
        dataset_root / filename
        for filename in SPLIT_FILES.values()
        if not (dataset_root / filename).is_file()
    ]
    images_dir = dataset_root / "Clean images"
    if not images_dir.is_dir():
        missing.append(images_dir)
    if missing:
        raise FileNotFoundError("Missing required dataset path(s): " + compact(str(path) for path in missing))


def main() -> int:
    args = parse_args()
    dataset_root = args.dataset_root.resolve()
    output_dir = args.output_dir.resolve()
    validate_layout(dataset_root)
    text_path, csv_path, verdict = audit(dataset_root, output_dir, args.phash_threshold)
    print(f"Audit complete: {verdict}")
    print(f"Text report: {text_path}")
    print(f"CSV report:  {csv_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
